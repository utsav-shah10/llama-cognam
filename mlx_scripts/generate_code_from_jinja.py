import subprocess
import os
import pandas as pd
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import requests


def parse_prompt(prompt_line):
    """
    Parse the prompt line to extract Task, Jinja Template, and Blended Template.
    Expected format: "Task: <task_description>, Jinja Template: <jinja_code>, Blended Template: "
    """
    # Use regex to extract the components
    pattern = r"Task*:\s*(.*?)\s*,\s*Jinja\s+Template*:\s*(.*?)\s*,\s*Blended\s+Template*:*"
    match = re.match(pattern, prompt_line.strip())
    
    if match:
        task = match.group(1).strip()
        jinja_template = match.group(2).strip()
        return task, jinja_template
    else:
        # Fallback: if pattern doesn't match, return the whole prompt as task
        return prompt_line.strip(), "", ""

def get_model_sheet_name(model_name, adapter_path, text_file):
    """
    Generate a sheet name based on model name and adapter path.
    """
    # Extract model name (last part after /)
    model_short = model_name.split('/')[0] if '/' in model_name else model_name
    
    # Extract adapter name (last part of path)
    adapter_short = os.path.basename(adapter_path) if adapter_path else "no_adapter"
    
    # Create sheet name (Excel sheet names have 31 char limit)
    sheet_name = f"{model_short}_{text_file}"
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]
    
    return sheet_name

class ExcelManager:
    """
    Manages Excel file operations with openpyxl.
    """
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.workbook = None
        self.load_or_create_workbook()
    
    def load_or_create_workbook(self):
        """Load existing workbook or create new one."""
        if os.path.exists(self.excel_file):
            try:
                self.workbook = load_workbook(self.excel_file)
                print(f"Loaded existing Excel file: '{self.excel_file}' with sheets: {self.workbook.sheetnames}")
            except Exception as e:
                print(f"Could not load existing Excel file: {e}")
                self.workbook = Workbook()
                # Remove default sheet
                if 'Sheet' in self.workbook.sheetnames:
                    self.workbook.remove(self.workbook['Sheet'])
        else:
            self.workbook = Workbook()
            # Remove default sheet
            if 'Sheet' in self.workbook.sheetnames:
                self.workbook.remove(self.workbook['Sheet'])
            print(f"Created new Excel workbook")
    
    def add_row(self, sheet_name, task, jinja_template, output_template, lint_output):
        """Add a new row to the specified sheet."""
        # Check if sheet exists and delete it if it does
        if sheet_name in self.workbook.sheetnames:
            self.workbook.remove(self.workbook[sheet_name])
            print(f"Deleted existing sheet: '{sheet_name}'")
        
        # Create new sheet
        worksheet = self.workbook.create_sheet(sheet_name)
        # Add headers with formatting
        headers = ['Task', 'Jinja Template', 'Output Template', 'Lint Output']
        worksheet.append(headers)
        
        # Format headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 20  # Task
        worksheet.column_dimensions['B'].width = 30  # Jinja Template
        worksheet.column_dimensions['C'].width = 30  # Output Template
        worksheet.column_dimensions['D'].width = 20  # Lint Template
        
        print(f"Created new sheet: '{sheet_name}'")
        
        # Add the new row
        worksheet.append([task, jinja_template, output_template, lint_output])
        
        # Format the new row
        row_num = worksheet.max_row
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, 5):  # 4 columns
            cell = worksheet.cell(row=row_num, column=col)
            cell.border = border
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical='top')
        
        print(f"Added row {row_num} to sheet '{sheet_name}'")
    
    def save_to_excel(self):
        """Save the workbook to Excel file."""
        try:
            self.workbook.save(self.excel_file)
            print(f"Saved Excel file: '{self.excel_file}' with sheets: {self.workbook.sheetnames}")
        except Exception as e:
            print(f"Error saving Excel file: {e}")
    
    def close(self):
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()

def generate_blended_code(
    input_prompts_file: str,
    excel_file: str,
    model_name: str = "mlx-community/Phi-3.5-mini-instruct-4bit",
    max_tokens: int = 750,
    adapter_path: str = "adapters_phi_2"
):
    
    # Create Excel directory if it doesn't exist
    excel_dir = os.path.dirname(excel_file)
    if excel_dir and not os.path.exists(excel_dir):
        os.makedirs(excel_dir)
        print(f"Created Excel directory: {excel_dir}")

    # Initialize Excel manager
    excel_manager = ExcelManager(excel_file)

    input_file_name = os.path.basename(input_prompts_file)

    # Generate sheet name for this model/adapter combination
    sheet_name = get_model_sheet_name(model_name, adapter_path, input_file_name)
    print(f"Using Excel sheet name: '{sheet_name}'")

    # Read all prompts from the input file
    with open(input_prompts_file, 'r', encoding='utf-8') as f:
        prompts = f.readlines()

    print(f"Found {len(prompts)} prompts in '{input_prompts_file}'")

    # Process first prompt to create the sheet (and delete if it exists)
    first_prompt_processed = False

    # Iterate through each prompt and generate code
    for i, prompt_line in enumerate(prompts):
        # Clean up the prompt line (remove leading/trailing whitespace, especially newlines)
        prompt = prompt_line.strip()

        # Parse the prompt to extract components
        task, jinja_template = parse_prompt(prompt)

        # reasoning prompt
        # final_prompt = f"Instruction: As an expert in code conversion, your task is to convert the provided Jinja template to its equivalent Blended template. Additionally, you must thoroughly explain the reasoning behind the specific conversion choices, demonstrating a deep understanding of Blended's strict type checking and scoping rules. Blended is a variant of Jinja with strict type checking and scoping rules. Do not include any other text, just provide translated code and below its reason. {prompt}"

        # without reasoning
        final_prompt = f"""You are an expert in Blended language template generation. Write blended template for following task which is compliant with Blended rules you know.
        Use provided jinja template as help in generating correct blended code.

        Task : {task}
        Jinja Template : {jinja_template}

        DO NOT OUTPUT ANY OTHER TEXT.
        """

        # Skip empty lines in the input file
        if not task:
            print(f"Skipping empty prompt line {i+1}.")
            continue

        print(f"\n--- Processing Prompt {i+1}/{len(prompts)} ---")
        print(f"Task: {task}")
        # print(f"Jinja Template: {jinja_template}")

        # Construct the command to run mlx_lm.generate
        command = [
            "python",
            "-m",
            "mlx_lm.generate",
            "--model", model_name,
            "--max-tokens", str(max_tokens),
            "--adapter-path", adapter_path,
            "--prompt", final_prompt 
        ]

        try:
            # Execute the command using subprocess.run
            process = subprocess.run(
                command,
                capture_output=True,
                text=True,
                check=True
            )

            # The full output from mlx_lm.generate includes metrics and the generated code.
            full_output = process.stdout.strip()

            # The generated code is typically after the "==========" separator.
            parts = full_output.split('==========')
            if len(parts) > 1:
                generated_code = parts[1].strip()
            else:
                generated_code = full_output.strip()

            # Remove the <|end|> token if the model appends it
            if generated_code.endswith("<|end|>"):
                generated_code = generated_code[:-len("<|end|>")].strip()


            # Lint output
            try:
                response = requests.post(
                    'http://10.5.0.97:8080/lint',
                    json={'code': generated_code},
                    timeout=5
                )
                lint_result = response.json().get('lint', 'Lint server error')
            except Exception as lint_err:
                lint_result = f"Linting error: {str(lint_err)}"
            

            # Handle sheet creation/deletion logic
            if not first_prompt_processed:
                # First prompt: this will delete existing sheet if it exists and create new one
                excel_manager.add_row(sheet_name, task, jinja_template, generated_code, lint_result)
                first_prompt_processed = True
            else:
                # Subsequent prompts: add to existing sheet
                worksheet = excel_manager.workbook[sheet_name]
                worksheet.append([task, jinja_template, generated_code, lint_result])
                
                # Format the new row
                row_num = worksheet.max_row
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for col in range(1, 5):  # 3 columns
                    cell = worksheet.cell(row=row_num, column=col)
                    cell.border = border
                    cell.alignment = cell.alignment.copy(wrap_text=True, vertical='top')
                
                print(f"Added row {row_num} to sheet '{sheet_name}'")

            print(f"Successfully processed prompt {i+1}")
            print(f"Generated Output Template (snippet): {generated_code[:200]}...")

        except subprocess.CalledProcessError as e:
            # Handle errors from the mlx_lm.generate command itself
            print(f"Error running mlx_lm.generate for prompt {i+1}:")
            print(f"Command: {' '.join(e.cmd)}")
            print(f"Return Code: {e.returncode}")
            print(f"Stderr: {e.stderr}")
            
            # Still save to Excel with error message
            error_message = f"ERROR: {e.stderr}"
            if not first_prompt_processed:
                excel_manager.add_row(sheet_name, task, jinja_template, error_message)
                first_prompt_processed = True
            else:
                worksheet = excel_manager.workbook[sheet_name]
                worksheet.append([task, jinja_template, error_message])
            
        except Exception as e:
            # Handle any other unexpected errors
            print(f"An unexpected error occurred for prompt {i+1}: {e}")
            error_message = f"ERROR: {str(e)}"
            if not first_prompt_processed:
                excel_manager.add_row(sheet_name, task, jinja_template, error_message)
                first_prompt_processed = True
            else:
                worksheet = excel_manager.workbook[sheet_name]
                worksheet.append([task, jinja_template, error_message])

    # Save all data to Excel file
    excel_manager.save_to_excel()
    excel_manager.close()
    print(f"All results saved to Excel file: '{excel_file}'")


# --- Example Usage ---
if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_code_from_jinja.py <input_txt_file> <output_excel_file>")
        sys.exit(1)

    input_file = sys.argv[1]
    excel_file = sys.argv[2]

    generate_blended_code(
        input_prompts_file=input_file,
        excel_file=excel_file,
        model_name="../model/qwen-tuned",
        adapter_path="../adapters_docs_new_tuned_finetune"
    )

    print("\n--- Script execution finished ---")