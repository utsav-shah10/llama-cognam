import pandas as pd
from pathlib import Path
from langchain_community.docstore.document import Document
from langchain_huggingface.embeddings import HuggingFaceEmbeddings
from langchain_community.vectorstores import FAISS
import os
import subprocess
from rich.console import Console
from rich.panel import Panel
from rich.prompt import Prompt
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side


MODEL_NAME = "Qwen/Qwen2.5-Coder-3B-Instruct"
os.environ["TOKENIZERS_PARALLELISM"] = "false" # has to do something with huggingface tokeniser library to avoid crashes and warnings. 

CSV_PATH = "../datasets/docs_and_code_grouped.csv"
FAISS_PATH = "faiss_index_docs_code" # this is where vector db is created
RESULTS_FILE = "../results_excel/rag_results_finetuned.xlsx"
ADAPTERS_PATH = "../adapters_docs_code_finetune"
FINETUNED_MODEL = False

console = Console() # for beautiful console (colors and design)

def load_and_chunk_csv(csv_path: str) -> list[Document]:

    """
        Role of this function is to create chunks of given documents and convert it to langchain doc format. 

        How chunks are created? 
        For our use case, we have dataset containing rule, qna pairs. So, one chunk contain all qna pairs belonging to one rule. 

        We did this to ensure that each chunk is unique and contain all info about one rule. 
    
        We created groups for code samples because if not done, all will go in one group and when retrieval is done, all code samples chunk will come as match and will be sent to LLM as context to given query. 

        Format of one chunk is : 
        Rule : Lower is a function. 
        
        Q: What is lower in blended?
        A: Lower is a builtin function in blended. 
    """
    df = pd.read_csv(csv_path)
    df.fillna("", inplace=True) # used when we didn't have any group for code samples, which we then introduced. 

    # print(df.head(10))

    grouped = df.groupby("rule") # groups are created based on column 'rule' in dataset
    documents = []

    for rule, group in grouped:
        qa_pairs = "\n".join(
            [f"Q: {row['question']}\nA: {row['answer']}" for _, row in group.iterrows()]
        )

        rule_text = "Code Example\n" if len(rule)==0 else rule
        chunk_text = f"Rule: {rule_text}\n\n{qa_pairs}"
        documents.append(Document(page_content=chunk_text, metadata={"rule": rule_text}))

    return documents


def create_faiss_index(documents, faiss_path="faiss_index"):

    """
        This is used to create a vector DB - we used FAISS vector db provided by facebook. 
        Each document is converted to vector embedding (using separate LLM) and then stored into vector DB. 
    """

    embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
    db = FAISS.from_documents(documents, embeddings)
    db.save_local(faiss_path)
    return db


def retrieve_docs(query, faiss_path="faiss_index", k=3):

    """
        This function is used to retrieve top k similar documents to query from our vector DB. 
        Query is converted to vector embedding and then similarity search is performed in DB to retrieve top k matches. 
        And then, these documents are passed as context to LLM to answer the query. 

        We can use different searching methods also like MMR (which ensures diversity in matches), but to keep it simple we used this default one. 

        We can also rank our retrieved documents using some another model and then pass top 'x' docs as context. 
    """

    embeddings = HuggingFaceEmbeddings(model_name="sentence-transformers/all-MiniLM-L6-v2")
    db = FAISS.load_local(faiss_path, embeddings, allow_dangerous_deserialization=True)
    return db.similarity_search(query, k=k)


def build_context(docs):
    return "\n\n".join([doc.page_content for doc in docs])


def generate_response(question, context):

    """
        This takes the question and retrieved context and answer the question based on context. If context is insufficient, it says, "insufficient info".
    """

    prompt = f"""\
        You are a precise assistant. Use the context below to answer the question.

        Use the context to answer the question as precisely as possible. If the context strongly implies the answer, you may infer it. If you are asked to write some code, write it based on context. Use your knowledge about blended. If nothing relevant is present, say: "Insufficient Information".


        Context:
        {context}

        Question:
        {question}

        Answer:"""

    try:

        base_command = [
                "python",
                "-m",
                "mlx_lm", 
                "generate",
                "--model", MODEL_NAME,
                "--prompt", prompt,
                "--max-tokens", "4096",
            ]
        
        if FINETUNED_MODEL:
            base_command.append("--adapter-path")
            base_command.append(ADAPTERS_PATH)

        # print(base_command)

        result = subprocess.run(
            base_command
            ,
            capture_output=True, text=True, check=True
        )
        output = result.stdout.strip()
        response = output.split("==========")[1]
        # print(rules_list)
        return response

    except Exception as e:
        print(f"⚠️ Error generating for question: {question}\n{e}")
        return []


def write_to_excel(sheet_name, qas):
    if Path(RESULTS_FILE).exists():
        wb = load_workbook(RESULTS_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(["Question", "Answer"])
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col in range(1, 3):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

    for question, answer in qas:
        ws.append([question, answer])
        row_num = ws.max_row
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col in range(1, 3):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border

    wb.save(RESULTS_FILE)


# Similar file mode as finetuning.

def file_mode(input_file):
    if not os.path.exists(input_file):
        print(f"❌ File not found: {input_file}")
        return

    with open(input_file, "r", encoding="utf-8") as f:
        questions = [line.strip() for line in f if line.strip()]

    qas = []
    for i, question in enumerate(questions):
        print(f"[{i+1}/{len(questions)}] Generating for: {question}")
        retrieved = retrieve_docs(question, faiss_path=FAISS_PATH)
        context = build_context(retrieved)
        response = generate_response(question, context)
        qas.append((question, response))


    sheet_name = Path(input_file).stem[:31]  # Excel sheet name max length is 31
    write_to_excel(sheet_name, qas)
    print(f"✅ Completed. Saved to sheet '{sheet_name}' in '{RESULTS_FILE}'")

# Similar chat mode as finetuning. 
def chat_rag_loop():


    session = []
    sheet_name = "rag_chat_session"

    console.print(Panel.fit("[bold green]Welcome to Blended RAG CLI[/bold green]\nType your question, or type 'exit' to quit.", title="🤖 RAG System"))

    while True:
        query = Prompt.ask("\n[bold cyan]Please enter a query[/bold cyan]")

        if query.lower() in ["exit", "quit"]:
            console.print("\n[bold red]Exiting RAG system. Goodbye![/bold red]")
            break

        console.print("[bold yellow]Retrieving best documents...[/bold yellow]")
        retrieved = retrieve_docs(query, faiss_path=FAISS_PATH)
        context = build_context(retrieved)

        console.print("[bold green]Generating response...[/bold green]")
        response = generate_response(query, context)

        # print(response)

        session.append((query, response))

        console.print(Panel.fit(response.strip(), title="📤 Answer", border_style="bold blue"))
    
    if session:
        write_to_excel(sheet_name, session)
        print(f"✅ Saved chat session to '{RESULTS_FILE}' in sheet '{sheet_name}'")


def main():

    """
        We take 3 arguments from user. 
        1. file name ofcourse
        2. --chat mode or file mode
        3. --f option which means, we want to use the fine tuned model. If option is used, adapters are used while generating the response, else, normal base model will be used. 
    
    """

    if len(sys.argv) < 2:
        print("Usage:")
        print("Use option --f to use fine tuned model with RAG.")
        print("  python rag.py --chat --f (optional)")
        print("  python rag.py <questions_txt_file> --f (optional)")
        return

    
    if len(sys.argv) == 3 and sys.argv[2] == "--f":
        print("Using finetuned model!")
        global FINETUNED_MODEL
        FINETUNED_MODEL = True

    # If vector db doesn't exist, we create it first and then rest of the things are done
    if not Path(FAISS_PATH).exists():
        console.print("[bold yellow]Building FAISS index...[/bold yellow]")
        docs = load_and_chunk_csv(CSV_PATH)
        create_faiss_index(docs, FAISS_PATH)

    if sys.argv[1] == "--chat":
        chat_rag_loop()
    else:
        file_mode(sys.argv[1])


if __name__ == "__main__":
    main()


