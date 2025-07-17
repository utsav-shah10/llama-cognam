# This file is used to run the fine tuned model on given tests. It is similar to generate_code_from_jinja.py file only. 

import subprocess
import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from pathlib import Path

MODEL_NAME = "Qwen/Qwen2.5-Coder-3B-Instruct"
ADAPTER_PATH = "../adapters_docs_code_finetune"
RESULTS_FILE = "../results_excel/qa_results.xlsx"


def generate_answer(question):

    """
        This function is used to generate the answer for a given question.
    """

    prompt = f"""You are an expert in Blended templating language. Answer the question in precise format.
    
    Question: {question}
    
    Answer:
    
    """
    try:
        result = subprocess.run(
            [
                "python", "-m", "mlx_lm.generate",
                "--model", MODEL_NAME,
                "--adapter-path", ADAPTER_PATH,
                "--prompt", prompt,
                "--max-tokens", "1024",
            ],
            capture_output=True, text=True, check=True
        )
        output = result.stdout.strip()
        if "==========" in output:
            answer = output.split("==========")[1].strip()
        else:
            answer = output.strip()

        if answer.endswith("<|end|>"):
            answer = answer[:-len("<|end|>")].strip()

        return answer
    except Exception as e:
        return f"⚠️ Error: {e}"


# this function writes the output to excel file in specified sheet name. Directly pasted from prev file.
def write_to_excel(sheet_name, qas):
    if Path(RESULTS_FILE).exists():
        wb = load_workbook(RESULTS_FILE)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(["Question", "Answer"])
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col in range(1, 3):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

    for question, answer in qas:
        ws.append([question, answer])
        row_num = ws.max_row
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col in range(1, 3):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border

    wb.save(RESULTS_FILE)


# Chat mode is also enabled in this to do continuous QnA like GPT. 
# Can be accessed using --chat option while running the file 
# All chat convo (que and ans) are stored into excel file to review afterwards if needed.

def chat_mode():
    print("🧠 Blended Chat (type 'exit' to quit)")
    session = []
    sheet_name = "chat_session"

    while True:
        try:
            question = input("You: ").strip()
            if question.lower() in ["exit", "quit"]:
                break
            if not question:
                continue

            print("💬 Thinking...\n")
            answer = generate_answer(question)
            print(f"\033[1mLLM:\033[0m {answer}\n")

            session.append((question, answer))
        except KeyboardInterrupt:
            break

    if session:
        write_to_excel(sheet_name, session)
        print(f"✅ Saved chat session to '{RESULTS_FILE}' in sheet '{sheet_name}'")


# File mode is directly giving the text file containing tasks as input and results will be saved in excel file. 

def file_mode(input_file):
    if not os.path.exists(input_file):
        print(f"❌ File not found: {input_file}")
        return

    with open(input_file, "r", encoding="utf-8") as f:
        questions = [line.strip() for line in f if line.strip()]

    qas = []
    for i, question in enumerate(questions):
        print(f"[{i+1}/{len(questions)}] Generating for: {question}")
        answer = generate_answer(question)
        qas.append((question, answer))

    sheet_name = Path(input_file).stem[:31]  # Excel sheet name max length is 31
    write_to_excel(sheet_name, qas)
    print(f"✅ Completed. Saved to sheet '{sheet_name}' in '{RESULTS_FILE}'")


def main():
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python run_eval.py --chat")
        print("  python run_eval.py <questions_txt_file>")
        return

    if sys.argv[1] == "--chat":
        chat_mode()
    else:
        file_mode(sys.argv[1])


if __name__ == "__main__":
    main()
